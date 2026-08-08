
"""
GVM Infra Developers - Telangana Construction & Renovation Website
================================================================
Features: Sign up / Login, Renovation & Build-New cost estimators
(residential + commercial, room-wise Telangana 2026 rates),
appointment booking (WhatsApp), Admin dashboard, Excel data storage.

Run:  pip install flask openpyxl werkzeug
      export ADMIN_PASSWORD=... SECRET_KEY=...   (see .env.example)
      python app.py
Admin login:  admin / (see ADMIN_PASSWORD env var)
"""

import os
import secrets
from datetime import datetime
from threading import Lock
from functools import wraps

from flask import (Flask, request, redirect, url_for, session,
                   render_template_string, send_file, flash)
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook

import security
import mailer

# ----------------------------- CONFIG ---------------------------------
COMPANY_NAME    = "GVM Infra Developers"
COMPANY_TAGLINE = "Building Telangana, One Home at a Time"
COMPANY_TAGLINE_TE = "మీ కల ఇల్లు – మా బాధ్యత"
PHONE           = "+91 8332899003"
WHATSAPP        = "918332899003"          # digits only, for wa.me links
SERVICE_AREA    = "All of Telangana"
# Company address shown on the site and used as Reply-To on customer email.
COMPANY_EMAIL   = os.environ.get("COMPANY_EMAIL", "vaibhavreddy@gvminfradevelopers.com")
ADMIN_USER      = "admin"
ADMIN_PASSWORD  = os.environ.get("ADMIN_PASSWORD", "dev-admin-change-me")
SECRET_KEY      = os.environ.get("SECRET_KEY", "dev-secret-change-me")
if ADMIN_PASSWORD == "dev-admin-change-me" or SECRET_KEY == "dev-secret-change-me":
    print("WARNING: ADMIN_PASSWORD/SECRET_KEY not set in environment — using insecure dev defaults.")
DATA_DIR        = os.environ.get("DATA_DIR", os.path.dirname(os.path.abspath(__file__)))
EXCEL_FILE      = os.path.join(DATA_DIR, "database.xlsx")

# ------------------- TELANGANA 2026 PRICES (Rs / sqft) ----------------
# Sources: NoBroker, Infralens, GharKaBudget, AECORD (July 2026 averages)
RENOVATION_PRICES = {
    "residential": {
        "Bedroom":        {"en": "Bedroom",        "te": "పడక గది",      "rate": 1050},
        "Bathroom":       {"en": "Bathroom",       "te": "స్నానాల గది",   "rate": 2200},
        "Kitchen":        {"en": "Kitchen (Modular)", "te": "వంటగది",    "rate": 2800},
        "Living Room":    {"en": "Living / Hall",  "te": "హాలు",          "rate": 1000},
        "Full House":     {"en": "Full House",     "te": "పూర్తి ఇల్లు",   "rate": 1250},
    },
    "commercial": {
        "Office Space":   {"en": "Office Space",   "te": "ఆఫీస్",         "rate": 1400},
        "Washroom":       {"en": "Washroom",       "te": "వాష్‌రూమ్",      "rate": 2500},
        "Pantry/Kitchen": {"en": "Pantry / Kitchen", "te": "పాంట్రీ",     "rate": 3000},
        "Retail/Shop":    {"en": "Retail / Shop Floor", "te": "షాపు",     "rate": 1600},
        "Full Property":  {"en": "Full Property",  "te": "పూర్తి భవనం",   "rate": 1550},
    },
}
BUILD_PRICES = {
    "residential": {
        "Basic":    {"rate": 1650, "desc": "Solid structure, standard fittings"},
        "Standard": {"rate": 1950, "desc": "Vitrified tiles, branded fittings"},
        "Premium":  {"rate": 2600, "desc": "Designer finishes, premium brands"},
    },
    "commercial": {
        "Basic":    {"rate": 1850, "desc": "Grey structure + basic finish"},
        "Standard": {"rate": 2200, "desc": "Ready-to-occupy commercial finish"},
        "Premium":  {"rate": 2900, "desc": "High-end showroom / office grade"},
    },
}

app = Flask(__name__)
app.secret_key = SECRET_KEY
xl_lock = Lock()

# Security headers, CSRF, hardened cookies, brute-force lockouts and the
# malicious-activity email alerts all live in security.py.
security.init_app(app)

# --------------------------- EXCEL BACKEND ----------------------------
SHEETS = {
    "Users":        ["ID", "Name", "Email", "Phone", "PasswordHash", "SignupDate"],
    "Estimates":    ["ID", "UserEmail", "UserName", "Service", "PropertyType",
                     "Details", "TotalSqft", "EstimatedCost", "Date"],
    "Appointments": ["ID", "Name", "Phone", "Email", "Service", "PreferredDate",
                     "PreferredTime", "Location", "Notes", "Status", "BookedOn"],
}

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        os.makedirs(DATA_DIR, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)
        for name, headers in SHEETS.items():
            ws = wb.create_sheet(name)
            ws.append(headers)
        wb.save(EXCEL_FILE)

def xl_append(sheet, row):
    with xl_lock:
        init_excel()  # self-heal if file was deleted
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet]
        # Neutralise spreadsheet formula injection: a value like ="=HYPERLINK(...)"
        # is harmless on the website but would execute when the workbook is opened.
        row = [ws.max_row] + [security.spreadsheet_safe(v) for v in row]  # auto ID
        ws.append(row)
        wb.save(EXCEL_FILE)
        return row[0]

def xl_rows(sheet):
    with xl_lock:
        init_excel()  # self-heal if file was deleted
        wb = load_workbook(EXCEL_FILE, read_only=True)
        ws = wb[sheet]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        return rows

def safe_float(v):
    """Parse user-supplied number safely; never negative, never crash."""
    try:
        return max(0.0, float(v))
    except (TypeError, ValueError):
        return 0.0

def form_key(key):
    """HTML-safe form field name for a room key."""
    return "sqft_" + key.replace(" ", "_").replace("/", "_")

def find_user(email):
    for r in xl_rows("Users"):
        if r[2] and r[2].lower() == email.lower():
            return {"id": r[0], "name": r[1], "email": r[2], "phone": r[3], "pw": r[4]}
    return None

init_excel()

# ----------------------------- HELPERS --------------------------------
def login_required(f):
    @wraps(f)
    def wrap(*a, **k):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*a, **k)
    return wrap

def admin_required(f):
    @wraps(f)
    def wrap(*a, **k):
        if not session.get("is_admin"):
            return redirect(url_for("admin_login"))
        return f(*a, **k)
    return wrap

def inr(n):
    """Indian number format: 12,34,567"""
    try:
        n = int(round(float(n)))
    except (TypeError, ValueError):
        return "0"
    s = str(n)
    if len(s) <= 3:
        return s
    last3, rest = s[-3:], s[:-3]
    parts = []
    while len(rest) > 2:
        parts.insert(0, rest[-2:])
        rest = rest[:-2]
    if rest:
        parts.insert(0, rest)
    return ",".join(parts + [last3])

app.jinja_env.filters["inr"] = inr
app.jinja_env.filters["ist"] = security._ist
app.jinja_env.globals["fkey"] = form_key

# ----------------------------- TEMPLATES ------------------------------
BASE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{ company }} | Construction & Renovation in Telangana</title>
<link rel="icon" type="image/png" href="{{ url_for('static', filename='favicon.png') }}">
<link rel="apple-touch-icon" href="{{ url_for('static', filename='favicon.png') }}">
<meta name="theme-color" content="#0B3B63">
<meta property="og:title" content="{{ company }} — Building Trust">
<meta property="og:description" content="Transparent per-sqft renovation and new-construction estimates across Telangana.">
<meta property="og:type" content="website">
<meta property="og:image" content="{{ url_for('static', filename='og-image.png', _external=True) }}">
<meta name="twitter:card" content="summary_large_image">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Big+Shoulders+Display:wght@600;700;800&family=IBM+Plex+Mono:wght@500;600&family=Noto+Sans:ital,wght@0,400;0,600;0,700;1,400&family=Noto+Sans+Telugu:wght@400;600&display=swap" rel="stylesheet">
<style>
:root{
  --blueprint:#0B3B63; --blueprint-d:#082C4C;
  --paper:#EEF3F8; --card:#fff;
  --chalk:#F6F3EA;
  --laterite:#A8431F; --laterite-bright:#C1502B;
  --ink:#16212C; --mut:#5C7086; --line:#CBD8E4;
}
*{margin:0;padding:0;box-sizing:border-box;}
html{scroll-behavior:smooth;}
body{background:var(--paper);color:var(--ink);min-height:100vh;display:flex;flex-direction:column;
     font-family:'Noto Sans','Noto Sans Telugu',system-ui,sans-serif;}
h1,h2{font-family:'Big Shoulders Display',sans-serif;font-weight:800;letter-spacing:.01em;text-transform:uppercase;}
.mono,.rate,.total,td.num,th.num,input[type=number]{font-family:'IBM Plex Mono',ui-monospace,monospace;}
a{color:inherit;}
:focus-visible{outline:3px solid var(--laterite-bright);outline-offset:2px;}
@media(prefers-reduced-motion:reduce){*{animation:none!important;transition:none!important;}}

header{background:var(--blueprint);color:var(--chalk);padding:16px 5%;display:flex;justify-content:space-between;
       align-items:center;flex-wrap:wrap;gap:10px;border-bottom:2px dashed rgba(246,243,234,.35);}
.logo{display:flex;align-items:center;text-decoration:none;flex-shrink:0;}
.logo-img{height:42px;width:auto;display:block;}
nav a{text-decoration:none;margin-left:20px;font-size:.85rem;font-weight:600;letter-spacing:.04em;text-transform:uppercase;
      color:var(--chalk);border-bottom:2px solid transparent;padding-bottom:2px;}
nav a:hover{border-color:var(--laterite-bright);}

.btn{display:inline-flex;align-items:center;gap:8px;background:var(--laterite);color:var(--chalk);font-weight:700;
    padding:13px 24px;border-radius:3px;text-decoration:none;border:none;cursor:pointer;font-size:.95rem;
    letter-spacing:.02em;transition:transform .15s,background .15s;}
.btn:hover{background:var(--laterite-bright);transform:translateY(-2px);}
.btn-outline{background:transparent;border:2px solid var(--blueprint);color:var(--blueprint);}
.btn-outline:hover{background:var(--blueprint);color:var(--chalk);}
.btn-wa{background:#1e7a4c;color:#fff;}
.btn-wa:hover{background:#186039;}

main{flex:1;}
.wrap{max-width:1050px;margin:0 auto;padding:40px 5%;}
h1{font-size:2.1rem;margin-bottom:10px;line-height:1.05;}
h2{font-size:1.3rem;margin:24px 0 14px;color:var(--blueprint);}
.te{color:var(--mut);font-size:.92rem;font-style:italic;}

.card{position:relative;background:var(--card);border:1px solid var(--line);border-radius:3px;padding:26px;}
.card::before,.card::after{content:"";position:absolute;width:10px;height:10px;border:2px solid var(--laterite-bright);
                            opacity:0;transition:opacity .2s;}
.card::before{top:-1px;left:-1px;border-right:none;border-bottom:none;}
.card::after{bottom:-1px;right:-1px;border-left:none;border-top:none;}
.grid2{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:22px;margin-top:22px;}
.opt{text-align:center;cursor:pointer;transition:transform .15s;text-decoration:none;color:var(--ink);}
.opt:hover{transform:translateY(-3px);}
.opt:hover::before,.opt:hover::after{opacity:1;}
.ic{width:44px;height:44px;margin:0 auto 10px;color:var(--laterite);}
.feature{text-align:center;}
.feature h3{font-size:1.05rem;margin-bottom:6px;}

.table-wrap{overflow-x:auto;margin-top:12px;border:1px solid var(--line);-webkit-overflow-scrolling:touch;}
table{width:100%;min-width:520px;border-collapse:collapse;background:#fff;}
.table-wrap table{margin-top:0;border:none;}
th,td{padding:10px 12px;border-bottom:1px solid var(--line);text-align:left;font-size:.9rem;white-space:nowrap;}
th{background:var(--blueprint);color:var(--chalk);font-size:.75rem;letter-spacing:.06em;text-transform:uppercase;
   font-weight:700;font-family:'Noto Sans',sans-serif;}
tr:hover td{background:#faf1ec;}
td .rate,td.rowcost{font-family:'IBM Plex Mono',monospace;}

input,select,textarea{width:100%;padding:11px;margin:6px 0 14px;border:1px solid var(--line);border-radius:3px;
                       font-size:1rem;font-family:'Noto Sans','Noto Sans Telugu',sans-serif;background:#fff;color:var(--ink);}
label{font-weight:700;font-size:.85rem;letter-spacing:.02em;text-transform:uppercase;color:var(--blueprint);}
.flash{background:#fdeee7;border:1px solid var(--laterite-bright);color:var(--laterite);padding:10px 14px;
      border-radius:3px;margin-bottom:14px;font-weight:600;}

.total{position:relative;background:var(--blueprint);color:var(--chalk);padding:20px 22px;border-radius:3px;
      font-size:1.3rem;font-weight:700;margin-top:14px;font-family:'IBM Plex Mono',monospace;}
.total::after{content:"EST.";position:absolute;top:-9px;right:14px;background:var(--laterite);color:var(--chalk);
             font-family:'Big Shoulders Display',sans-serif;font-size:.65rem;letter-spacing:.1em;padding:2px 8px;}
.total span{color:var(--laterite-bright);}

footer{background:var(--blueprint);color:var(--chalk);text-align:center;padding:24px 20px;font-size:.85rem;
      border-top:2px dashed rgba(246,243,234,.35);}
footer a{color:var(--laterite-bright);text-decoration:none;}
footer a:hover{text-decoration:underline;}
.footer-mark{height:44px;width:auto;display:block;margin:0 auto 12px;opacity:.9;}
.fbr{display:none;}
@media(max-width:600px){.fbr{display:inline;}}

.hero{background:var(--blueprint);color:var(--chalk);padding:0;overflow:hidden;position:relative;}
.hero-inner{max-width:1050px;margin:0 auto;padding:56px 5% 44px;display:grid;grid-template-columns:1.1fr 1fr;gap:30px;align-items:center;}
.hero h1{font-size:2.7rem;}
.hero p{margin:14px 0 26px;font-size:1.05rem;color:#cddcea;max-width:46ch;}
.hero-cta{display:flex;gap:12px;flex-wrap:wrap;}
.blueprint-svg{width:100%;height:auto;}
.blueprint-svg text{font-family:'IBM Plex Mono',monospace;}
.callout-num{font-family:'IBM Plex Mono',monospace;font-weight:600;}

.badge{display:inline-block;background:#e7eef3;color:var(--blueprint);padding:3px 10px;border-radius:3px;
      font-size:.78rem;font-weight:700;letter-spacing:.03em;text-transform:uppercase;}
.badge.p{background:#fdeee7;color:var(--laterite);}
.stat-num{font-family:'IBM Plex Mono',monospace;font-size:1.6rem;font-weight:700;color:var(--blueprint);margin:0;}
.testimonial .stars{color:var(--laterite);font-family:'IBM Plex Mono',monospace;letter-spacing:.1em;font-size:1rem;}
.testimonial p.quote{margin:10px 0 14px;font-style:italic;}
.testimonial .who{font-size:.85rem;color:var(--mut);}
@media(max-width:760px){.hero-inner{grid-template-columns:1fr;}.hero h1{font-size:2rem;}}
@media(max-width:600px){
  nav{width:100%;justify-content:flex-start;}
  nav a{margin-left:0;margin-right:16px;}
}
@media(max-width:480px){
  .logo-img{height:34px;}
  .footer-mark{height:36px;}
  .wrap{padding:28px 5%;}
  .card{padding:18px;}
  .hero-inner{padding:40px 5% 32px;}
  .grid2{grid-template-columns:1fr;gap:16px;}
  .total{font-size:1.1rem;padding:16px 18px;}
}
</style>
</head>
<body>
<header>
  <a class="logo" href="{{ url_for('home') }}" aria-label="{{ company }} — home">
    <img class="logo-img" src="{{ url_for('static', filename='logo.png') }}"
         alt="{{ company }} — Building Trust" width="880" height="136">
  </a>
  <nav>
    {% if session.get('user') %}
      <a href="{{ url_for('dashboard') }}">Home</a>
      <a href="{{ url_for('appointment') }}">Book Visit</a>
      <a href="{{ url_for('logout') }}">Logout ({{ session['user']['name'].split()[0] if session['user']['name'].split() else 'User' }})</a>
    {% elif session.get('is_admin') %}
      <a href="{{ url_for('admin') }}">Admin</a><a href="{{ url_for('logout') }}">Logout</a>
    {% else %}
      <a href="{{ url_for('login') }}">Login</a><a href="{{ url_for('signup') }}">Sign Up</a>
    {% endif %}
  </nav>
</header>
<main>
{% with msgs = get_flashed_messages() %}{% if msgs %}<div class="wrap" style="padding-bottom:0">
{% for m in msgs %}<div class="flash">{{ m }}</div>{% endfor %}</div>{% endif %}{% endwith %}
{{ body }}
</main>
<footer>
<img class="footer-mark" src="{{ url_for('static', filename='logo-mark.png') }}" alt="" aria-hidden="true">
© 2026 {{ company }} · {{ area }}<br class="fbr">
<a href="tel:{{ phone|replace(' ','') }}">{{ phone }}</a> ·
{% if email %}<a href="mailto:{{ email }}">{{ email }}</a> · {% endif %}
<a href="https://wa.me/{{ wa }}">WhatsApp</a></footer>
</body></html>
"""

def page(body_tpl, **ctx):
    from markupsafe import Markup
    body = Markup(render_template_string(body_tpl, **ctx))  # already escaped by Jinja
    return render_template_string(BASE, body=body, company=COMPANY_NAME,
                                  phone=PHONE, wa=WHATSAPP, area=SERVICE_AREA,
                                  email=COMPANY_EMAIL)

# ------------------------------ ROUTES --------------------------------
@app.route("/")
def home():
    return page("""
<div class="hero">
  <div class="hero-inner">
    <div>
      <h1>{{ cname }}</h1>
      <p>{{ tagline }} · <b>{{ tagline_te }}</b><br>
      Every quote is measured, not guessed — transparent per-sqft rates for renovation
      and new construction across Telangana.</p>
      <div class="hero-cta">
        <a class="btn" href="{{ url_for('signup') }}">Get Instant Estimate →</a>
        <a class="btn btn-wa" href="https://wa.me/{{ wa }}?text=Hi, I want a construction estimate">WhatsApp Us</a>
      </div>
    </div>
    <svg class="blueprint-svg" viewBox="0 0 360 230" role="img" aria-label="Blueprint elevation of a house with a sample estimate of 1,850 square feet">
      <path d="M40 170 L40 95 L180 25 L320 95 L320 170 Z" fill="none" stroke="var(--chalk)" stroke-width="2"/>
      <rect x="160" y="128" width="40" height="42" fill="none" stroke="var(--chalk)" stroke-width="1.5"/>
      <rect x="70"  y="108" width="30" height="30" fill="none" stroke="var(--chalk)" stroke-width="1.4"/>
      <rect x="260" y="108" width="30" height="30" fill="none" stroke="var(--chalk)" stroke-width="1.4"/>
      <line x1="70" y1="123" x2="100" y2="123" stroke="var(--chalk)" stroke-width="1"/>
      <line x1="260" y1="123" x2="290" y2="123" stroke="var(--chalk)" stroke-width="1"/>
      <line x1="40" y1="190" x2="320" y2="190" stroke="var(--chalk)" stroke-width="1" stroke-dasharray="4 3" opacity=".6"/>
      <line x1="40" y1="182" x2="40" y2="198" stroke="var(--chalk)" stroke-width="1" opacity=".6"/>
      <line x1="320" y1="182" x2="320" y2="198" stroke="var(--chalk)" stroke-width="1" opacity=".6"/>
      <text x="180" y="212" fill="var(--chalk)" font-size="11" text-anchor="middle" opacity=".8">42 FT WIDE</text>
      <line x1="340" y1="25" x2="340" y2="170" stroke="var(--laterite-bright)" stroke-width="1.5" stroke-dasharray="4 3"/>
      <line x1="332" y1="25" x2="348" y2="25" stroke="var(--laterite-bright)" stroke-width="1.5"/>
      <line x1="332" y1="170" x2="348" y2="170" stroke="var(--laterite-bright)" stroke-width="1.5"/>
      <g transform="translate(180,90)">
        <rect x="-58" y="-20" width="116" height="40" fill="var(--laterite)" rx="2"/>
        <text x="0" y="-2" class="callout-num" fill="var(--chalk)" font-size="16" text-anchor="middle" data-count-to="1850">0</text>
        <text x="0" y="14" fill="var(--chalk)" font-size="9" text-anchor="middle" opacity=".85">SQ FT SAMPLE ESTIMATE</text>
      </g>
    </svg>
  </div>
</div>
<div class="wrap">
  <h2 style="text-align:center">How the estimate is built / అంచనా ఎలా తయారవుతుంది</h2>
  <div class="grid2">
    <div class="card feature">
      <div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round"><path d="M6 4h10M6 8h8M6 4c4.5 0 4.5 4.5 0 4.5H6L15 18"/></svg></div>
      <h3>Room-wise, per-sqft</h3><p class="te">Every rate is published up front — bedroom, bathroom, kitchen, hall. No hidden line items after the site visit.</p>
    </div>
    <div class="card feature">
      <div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"><path d="M3 16a9 9 0 0 1 18 0"/><path d="M2 16h20"/><path d="M12 7V5"/></svg></div>
      <h3>20+ Years on Telangana Sites</h3><p class="te">Homes, apartments and commercial builds across the state — rates tuned to today's local market, not a national average.</p>
    </div>
    <div class="card feature">
      <div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="5" width="18" height="16" rx="1"/><path d="M3 9h18M8 3v4M16 3v4"/><path d="m8.5 14 2 2 4-4"/></svg></div>
      <h3>Free Site Measurement</h3><p class="te">Book a visit online — we measure on site and turn your estimate into an exact quote at no charge.</p>
    </div>
  </div>
</div>
<div class="wrap">
  <h2 style="text-align:center">What Customers Say / మా వినియోగదారుల మాటల్లో</h2>
  <p class="te" style="text-align:center;margin-top:-6px">Sample reviews shown below — replace with real customer testimonials.</p>
  <section id="testimonials" class="grid2" aria-label="Customer testimonials">
    <div class="card testimonial">
      <div class="stars" aria-label="5 out of 5 stars">★★★★★</div>
      <p class="quote">"Room-wise pricing meant zero surprises. Our kitchen renovation in Hyderabad came in exactly at quote."</p>
      <p class="who"><b>Priya R.</b> · Hyderabad</p>
    </div>
    <div class="card testimonial">
      <div class="stars" aria-label="5 out of 5 stars">★★★★★</div>
      <p class="quote">"Booked the free site visit through the app and had a firm quote for our new house in two days."</p>
      <p class="who"><b>Srinivas M.</b> · Warangal</p>
    </div>
    <div class="card testimonial">
      <div class="stars" aria-label="4 out of 5 stars">★★★★☆</div>
      <p class="quote">"Good communication throughout our shop renovation. Telugu support made it easy for my father to follow along too."</p>
      <p class="who"><b>Anitha K.</b> · Karimnagar</p>
    </div>
    <div class="card testimonial">
      <div class="stars" aria-label="5 out of 5 stars">★★★★★</div>
      <p class="quote">"Transparent per-sqft rates for our office build-out — no hidden line items after the site visit like other contractors quoted."</p>
      <p class="who"><b>Rahul D.</b> · Nizamabad</p>
    </div>
    <div class="card testimonial">
      <div class="stars" aria-label="5 out of 5 stars">★★★★★</div>
      <p class="quote">"WhatsApp booking was so convenient. They measured our full house and matched the online estimate almost exactly."</p>
      <p class="who"><b>Lakshmi V.</b> · Khammam</p>
    </div>
  </section>
</div>
<script>
(function(){
  if (window.matchMedia && window.matchMedia('(prefers-reduced-motion: reduce)').matches) return;
  var el = document.querySelector('[data-count-to]');
  if (!el) return;
  var target = parseInt(el.getAttribute('data-count-to'), 10), start = null, dur = 900;
  function step(ts){
    if (!start) start = ts;
    var p = Math.min(1, (ts - start) / dur);
    el.textContent = Math.round(p * target).toLocaleString('en-IN');
    if (p < 1) requestAnimationFrame(step);
  }
  requestAnimationFrame(step);
})();
</script>""", cname=COMPANY_NAME, tagline=COMPANY_TAGLINE, tagline_te=COMPANY_TAGLINE_TE, wa=WHATSAPP)

@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        name  = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip().lower()
        phone = request.form.get("phone", "").strip()
        pw    = request.form.get("password", "")
        if len(name) > 80 or len(email) > 120 or len(phone) > 20 or len(pw) > 200:
            flash("Please shorten your details and try again.")
            return redirect(url_for("signup"))
        if not name or "@" not in email or len(phone) < 10 or len(pw) < 8:
            flash("Please fill all fields correctly (phone: 10+ digits, password: 8+ characters).")
            return redirect(url_for("signup"))
        if find_user(email):
            flash("An account with this email already exists. Please login.")
            return redirect(url_for("login"))
        # Counted only for real account creations, so a customer fumbling the
        # form a few times never gets rate-limited.
        if not security.signup_allowed():
            flash("Too many accounts created from this connection. Please try again later.")
            return redirect(url_for("signup"))
        xl_append("Users", [name, email, phone, generate_password_hash(pw, method="pbkdf2:sha256"),
                            datetime.now().strftime("%Y-%m-%d %H:%M")])
        session["user"] = {"name": name, "email": email, "phone": phone}
        # Greeting email. Sent on a background thread, so a slow or misconfigured
        # SMTP server can never delay or fail the signup itself.
        mailer.send_welcome(email, name)
        return redirect(url_for("dashboard"))
    return page("""
<div class="wrap" style="max-width:460px"><div class="card">
<h1>Sign Up</h1><p class="te">ఖాతా సృష్టించండి — free instant estimates</p>
<form method="post">
<input type="hidden" name="_csrf" value="{{ csrf_token() }}">
<label>Full Name / పేరు</label><input name="name" required>
<label>Email</label><input type="email" name="email" required>
<label>Phone / ఫోన్</label><input name="phone" pattern="[0-9+ \\-]{10,15}" maxlength="20" required>
<label>Password</label><input type="password" name="password" minlength="6" required>
<button class="btn" style="width:100%">Create Account</button>
</form>
<p style="margin-top:12px" class="te">Already have an account? <a href="{{ url_for('login') }}">Login</a></p>
</div></div>""")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        if security.is_locked("loginfail"):
            flash("Too many failed attempts. Please wait a few minutes and try again.")
            return redirect(url_for("login"))
        u = find_user(email)
        if u and check_password_hash(u["pw"], request.form.get("password", "")):
            session["user"] = {"name": u["name"], "email": u["email"], "phone": u["phone"]}
            return redirect(url_for("dashboard"))
        security.login_failed(email)
        flash("Invalid email or password.")
    return page("""
<div class="wrap" style="max-width:460px"><div class="card">
<h1>Login</h1><p class="te">లాగిన్ చేయండి</p>
<form method="post">
<input type="hidden" name="_csrf" value="{{ csrf_token() }}">
<label>Email</label><input type="email" name="email" required>
<label>Password</label><input type="password" name="password" required>
<button class="btn" style="width:100%">Login</button>
</form>
<p style="margin-top:12px" class="te">New here? <a href="{{ url_for('signup') }}">Sign Up</a> ·
<a href="{{ url_for('admin_login') }}" style="color:#aaa">Admin</a></p>
</div></div>""")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))

@app.route("/dashboard")
@login_required
def dashboard():
    return page("""
<div class="wrap">
<h1>Welcome, {{ session['user']['name'] }}</h1>
<p class="te">మీకు ఏ సేవ కావాలి? — What would you like to do?</p>
<div class="grid2">
  <a class="card opt" href="{{ url_for('renovation') }}">
    <div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"><path d="M14 6l4 4-7 7-4-4z"/><path d="M4 20l3-3"/></svg></div>
    <h2>Option 1: Renovation</h2>
    <p class="te">పునరుద్ధరణ — Room-wise renovation cost for house or commercial property. Today's Telangana rates per sqft.</p>
  </a>
  <a class="card opt" href="{{ url_for('build_new') }}">
    <div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"><path d="M4 11 12 4l8 7"/><path d="M6 10v10h12V10"/><path d="M10 20v-6h4v6"/></svg></div>
    <h2>Option 2: Build New</h2>
    <p class="te">కొత్త నిర్మాణం — New residential or commercial construction cost estimate per sqft.</p>
  </a>
</div>
<div style="text-align:center;margin-top:30px">
  <a class="btn btn-wa" href="{{ url_for('appointment') }}">Schedule an Appointment / అపాయింట్‌మెంట్ బుక్ చేయండి</a>
</div>
</div>""")

# ------------------------- RENOVATION FLOW -----------------------------
@app.route("/renovation")
@login_required
def renovation():
    ptype = request.args.get("type")
    if ptype not in ("residential", "commercial"):
        return page("""
<div class="wrap">
<h1>Renovation</h1><p class="te">Is the renovation for a House or a Commercial property?</p>
<div class="grid2">
  <a class="card opt" href="{{ url_for('renovation', type='residential') }}">
    <div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"><path d="M4 11 12 4l8 7"/><path d="M6 10v10h12V10"/><path d="M10 20v-6h4v6"/></svg></div><h2>House / ఇల్లు</h2><p class="te">Bedroom, bathroom, kitchen, hall — room-wise rates.</p></a>
  <a class="card opt" href="{{ url_for('renovation', type='commercial') }}">
    <div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"><rect x="5" y="3" width="14" height="18"/><path d="M9 7h2M13 7h2M9 11h2M13 11h2M9 15h2M13 15h2"/></svg></div><h2>Commercial / వాణిజ్య</h2><p class="te">Office, shop, washroom, pantry — commercial rates.</p></a>
</div></div>""")
    rooms = RENOVATION_PRICES[ptype]
    return page("""
<div class="wrap">
<h1>{{ 'House' if ptype=='residential' else 'Commercial' }} Renovation — Telangana Rates (July 2026)</h1>
<p class="te">Enter area in sqft for each room you want to renovate. మీరు పునరుద్ధరించాలనుకునే గదుల విస్తీర్ణం నమోదు చేయండి.</p>
<form method="post" action="{{ url_for('save_estimate') }}" oninput="calc()">
<input type="hidden" name="_csrf" value="{{ csrf_token() }}">
<input type="hidden" name="service" value="Renovation">
<input type="hidden" name="ptype" value="{{ ptype }}">
<div class="table-wrap"><table>
<tr><th>Room / గది</th><th>Rate (₹/sqft)</th><th>Area (sqft)</th><th>Cost (₹)</th></tr>
{% for key, r in rooms.items() %}
<tr>
  <td><b>{{ r.en }}</b><br><span class="te">{{ r.te }}</span></td>
  <td>₹{{ r.rate }}</td>
  <td><input type="number" min="0" step="1" name="{{ fkey(key) }}" data-rate="{{ r.rate }}" class="sq" style="margin:0;max-width:130px" placeholder="0"></td>
  <td class="rowcost">₹0</td>
</tr>
{% endfor %}
</table></div>
<div class="total">Estimated Total / మొత్తం అంచనా: <span id="tot">₹0</span></div>
<p class="te" style="margin-top:8px">* Indicative estimate at current Telangana market rates. Final quote after free site visit.</p>
<button class="btn" style="margin-top:14px">Save Estimate</button>
<a class="btn btn-wa" style="margin-top:14px" href="{{ url_for('appointment') }}">Book Free Site Visit</a>
</form>
</div>
<script>
function calc(){let t=0;document.querySelectorAll('.sq').forEach(i=>{
 const c=(parseFloat(i.value)||0)*parseFloat(i.dataset.rate);t+=c;
 i.closest('tr').querySelector('.rowcost').textContent='₹'+c.toLocaleString('en-IN');});
 document.getElementById('tot').textContent='₹'+t.toLocaleString('en-IN');}
</script>""", ptype=ptype, rooms=rooms)

# -------------------------- BUILD NEW FLOW -----------------------------
@app.route("/build-new")
@login_required
def build_new():
    ptype = request.args.get("type")
    if ptype not in ("residential", "commercial"):
        return page("""
<div class="wrap">
<h1>Build New</h1><p class="te">What do you want to build? మీరు ఏమి నిర్మించాలనుకుంటున్నారు?</p>
<div class="grid2">
  <a class="card opt" href="{{ url_for('build_new', type='residential') }}">
    <div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"><path d="M4 11 12 4l8 7"/><path d="M6 10v10h12V10"/><path d="M10 20v-6h4v6"/></svg></div><h2>New Residential</h2><p class="te">Independent house, villa, duplex — full construction.</p></a>
  <a class="card opt" href="{{ url_for('build_new', type='commercial') }}">
    <div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"><rect x="5" y="3" width="14" height="18"/><path d="M9 7h2M13 7h2M9 11h2M13 11h2M9 15h2M13 15h2"/></svg></div><h2>New Commercial</h2><p class="te">Shops, offices, function halls, warehouses.</p></a>
</div></div>""")
    tiers = BUILD_PRICES[ptype]
    rooms = RENOVATION_PRICES[ptype]
    return page("""
<div class="wrap">
<h1>New {{ 'Residential' if ptype=='residential' else 'Commercial' }} Construction — Telangana Rates (July 2026)</h1>
<form method="post" action="{{ url_for('save_estimate') }}" oninput="calc()">
<input type="hidden" name="_csrf" value="{{ csrf_token() }}">
<input type="hidden" name="service" value="Build New">
<input type="hidden" name="ptype" value="{{ ptype }}">
<div class="card">
<label>Total Built-up Area (sqft) / మొత్తం విస్తీర్ణం</label>
<input type="number" id="area" name="sqft_total" min="0" required placeholder="e.g. 1800">
<label>Quality Package</label>
<select id="tier" name="tier">
{% for name, t in tiers.items() %}
<option value="{{ name }}" data-rate="{{ t.rate }}">{{ name }} — ₹{{ t.rate }}/sqft ({{ t.desc }})</option>
{% endfor %}
</select>
<div class="total">Estimated Total / మొత్తం అంచనా: <span id="tot">₹0</span></div>
</div>
<h2>Room-wise finishing rates (same as renovation rates)</h2>
<div class="table-wrap"><table>
<tr><th>Room / గది</th><th>Rate (₹/sqft)</th></tr>
{% for key, r in rooms.items() %}<tr><td>{{ r.en }} <span class="te">({{ r.te }})</span></td><td>₹{{ r.rate }}</td></tr>{% endfor %}
</table></div>
<p class="te" style="margin-top:8px">* Package covers full structure + finishing. Final quote after free site visit & plan review.</p>
<button class="btn" style="margin-top:14px">Save Estimate</button>
<a class="btn btn-wa" style="margin-top:14px" href="{{ url_for('appointment') }}">Book Free Site Visit</a>
</form>
</div>
<script>
function calc(){const a=parseFloat(document.getElementById('area').value)||0;
 const r=parseFloat(document.getElementById('tier').selectedOptions[0].dataset.rate);
 document.getElementById('tot').textContent='₹'+(a*r).toLocaleString('en-IN');}
</script>""", ptype=ptype, tiers=tiers, rooms=rooms)

# --------------------------- SAVE ESTIMATE -----------------------------
@app.route("/save-estimate", methods=["POST"])
@login_required
def save_estimate():
    service = request.form.get("service", "")
    ptype   = request.form.get("ptype", "")
    if service not in ("Renovation", "Build New") or ptype not in ("residential", "commercial"):
        flash("Something went wrong — please try again.")
        return redirect(url_for("dashboard"))
    details, total_sqft, total_cost = [], 0, 0
    if service == "Renovation":
        for key, r in RENOVATION_PRICES[ptype].items():
            sq = safe_float(request.form.get(form_key(key)))
            if sq > 0:
                cost = sq * r["rate"]
                details.append(f"{r['en']}: {int(sq)} sqft @ ₹{r['rate']} = ₹{inr(cost)}")
                total_sqft += sq; total_cost += cost
    else:
        sq   = safe_float(request.form.get("sqft_total"))
        tier = request.form.get("tier", "Standard")
        if tier not in BUILD_PRICES[ptype]:
            tier = "Standard"
        rate = BUILD_PRICES[ptype][tier]["rate"]
        total_sqft, total_cost = sq, sq * rate
        details.append(f"{tier} package: {int(sq)} sqft @ ₹{rate}/sqft")
    if total_sqft == 0:
        flash("Please enter at least one area in sqft.")
        # Only bounce back to our own pages — request.referrer is attacker-supplied.
        back = url_for("renovation" if service == "Renovation" else "build_new",
                       type=ptype)
        return redirect(back)
    u = session["user"]
    xl_append("Estimates", [u["email"], u["name"], service, ptype.title(),
                            "; ".join(details), int(total_sqft), int(total_cost),
                            datetime.now().strftime("%Y-%m-%d %H:%M")])
    return page("""
<div class="wrap" style="max-width:640px"><div class="card" style="text-align:center">
<div class="ic" style="width:56px;height:56px"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="9"/><path d="m8 12.5 2.5 2.5L16 9"/></svg></div>
<h1>Estimate Saved!</h1>
<p class="te">{{ service }} ({{ ptype }}) — {{ sq }} sqft</p>
<div class="total">Estimated Cost: <span>₹{{ cost|inr }}</span></div>
<p style="margin:16px 0" class="te">Our team will review your estimate. Book a free site visit for an exact quote!</p>
<a class="btn btn-wa" href="{{ url_for('appointment') }}">Book Free Site Visit</a>
<a class="btn btn-outline" href="{{ url_for('dashboard') }}">← Back to Home</a>
</div></div>""", service=service, ptype=ptype.title(), sq=int(total_sqft), cost=total_cost)

# --------------------------- APPOINTMENTS ------------------------------
@app.route("/appointment", methods=["GET", "POST"])
@login_required
def appointment():
    u = session["user"]
    if request.method == "POST":
        f = request.form
        name, phone = f.get("name", "").strip(), f.get("phone", "").strip()
        loc, date_ = f.get("location", "").strip(), f.get("date", "").strip()
        if not (name and phone and loc and date_):
            flash("Please fill in name, phone, date and location.")
            return redirect(url_for("appointment"))
        xl_append("Appointments", [name, phone, u["email"], f.get("service", ""),
                                   date_, f.get("time", ""), loc,
                                   f.get("notes", ""), "Pending",
                                   datetime.now().strftime("%Y-%m-%d %H:%M")])
        wa_msg = (f"Hi {COMPANY_NAME}, I booked an appointment. Name: {name}, "
                  f"Service: {f.get('service','')}, Date: {date_} {f.get('time','')}, Location: {loc}")
        return page("""
<div class="wrap" style="max-width:640px"><div class="card" style="text-align:center">
<div class="ic" style="width:56px;height:56px"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="5" width="18" height="16" rx="1"/><path d="M3 9h18M8 3v4M16 3v4"/><path d="m8.5 14 2 2 4-4"/></svg></div><h1>Appointment Booked!</h1>
<p class="te">మీ అపాయింట్‌మెంట్ నమోదైంది. We will call you at {{ phone }} to confirm.</p>
<a class="btn btn-wa" href="https://wa.me/{{ wa }}?text={{ msg|urlencode }}">Confirm on WhatsApp</a>
<a class="btn btn-outline" href="{{ url_for('dashboard') }}">← Back to Home</a>
</div></div>""", phone=request.form["phone"], wa=WHATSAPP, msg=wa_msg)
    return page("""
<div class="wrap" style="max-width:560px"><div class="card">
<h1>Schedule an Appointment</h1>
<p class="te">ఉచిత సైట్ విజిట్ — Free site visit & exact quotation</p>
<form method="post">
<input type="hidden" name="_csrf" value="{{ csrf_token() }}">
<label>Name</label><input name="name" value="{{ session['user']['name'] }}" required>
<label>Phone</label><input name="phone" value="{{ session['user']['phone'] }}" required>
<label>Service</label>
<select name="service"><option>Renovation - House</option><option>Renovation - Commercial</option>
<option>New Residential Construction</option><option>New Commercial Construction</option><option>Other / Consultation</option></select>
<label>Preferred Date</label><input type="date" name="date" min="{{ today }}" required>
<label>Preferred Time</label>
<select name="time"><option>Morning (9 AM - 12 PM)</option><option>Afternoon (12 - 4 PM)</option><option>Evening (4 - 7 PM)</option></select>
<label>Site Location (Village/City, District)</label><input name="location" placeholder="e.g. Karimnagar" required>
<label>Notes (optional)</label><textarea name="notes" rows="2"></textarea>
<button class="btn" style="width:100%">Book Appointment / బుక్ చేయండి</button>
</form></div></div>""", today=datetime.now().strftime("%Y-%m-%d"))

# --------------------------- ERROR PAGES -------------------------------
@app.errorhandler(404)
def not_found(e):
    return page("""<div class="wrap" style="text-align:center"><h1>Page not found</h1>
<a class="btn" href="{{ url_for('home') }}">← Go Home</a></div>"""), 404

@app.errorhandler(500)
def server_error(e):
    return page("""<div class="wrap" style="text-align:center"><h1>Something went wrong</h1>
<p class="te">Please try again or call us at """ + PHONE + """</p>
<a class="btn" href="{{ url_for('home') }}">← Go Home</a></div>"""), 500

# ------------------------------ ADMIN ----------------------------------
@app.route("/admin-login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if security.is_locked("adminfail"):
            flash("Too many failed attempts. This connection is locked for 15 minutes.")
            return redirect(url_for("admin_login"))
        user = request.form.get("username", "")
        pw   = request.form.get("password", "")
        # compare_digest keeps the check constant-time so the password can't be
        # recovered a character at a time by timing the response.
        ok = (secrets.compare_digest(user, ADMIN_USER)
              and secrets.compare_digest(pw, ADMIN_PASSWORD))
        if ok:
            session.clear()
            session["is_admin"] = True
            session.permanent = True
            security.admin_login_ok()
            return redirect(url_for("admin"))
        security.admin_login_failed(user)
        flash("Wrong admin credentials.")
    return page("""
<div class="wrap" style="max-width:420px"><div class="card">
<h1>🔐 Admin Login</h1>
<form method="post">
<input type="hidden" name="_csrf" value="{{ csrf_token() }}">
<label>Username</label><input name="username" required>
<label>Password</label><input type="password" name="password" required>
<button class="btn" style="width:100%">Login</button>
</form></div></div>""")

@app.route("/admin")
@admin_required
def admin():
    users, ests, appts = xl_rows("Users"), xl_rows("Estimates"), xl_rows("Appointments")
    total_value = sum((r[7] or 0) for r in ests)
    events = security.recent_events(40)
    return page("""
<div class="wrap">
<h1>🛠️ Admin Dashboard</h1>
<div class="grid2" style="grid-template-columns:repeat(auto-fit,minmax(180px,1fr))">
  <div class="card" style="text-align:center"><div class="stat-num">{{ users|length }}</div><p class="te">Customers</p></div>
  <div class="card" style="text-align:center"><div class="stat-num">{{ ests|length }}</div><p class="te">Estimates</p></div>
  <div class="card" style="text-align:center"><div class="stat-num">{{ appts|length }}</div><p class="te">Appointments</p></div>
  <div class="card" style="text-align:center"><div class="stat-num">₹{{ tv|inr }}</div><p class="te">Pipeline Value</p></div>
</div>
<div style="margin:18px 0"><a class="btn" href="{{ url_for('download_excel') }}">⬇️ Download Excel (all data)</a></div>

<h2>📅 Appointments</h2>
<div class="table-wrap"><table><tr><th>ID</th><th>Name</th><th>Phone</th><th>Service</th><th>Date</th><th>Time</th><th>Location</th><th>Status</th></tr>
{% for a in appts|reverse %}<tr><td>{{ a[0] }}</td><td>{{ a[1] }}</td><td><a href="tel:{{ a[2] }}">{{ a[2] }}</a></td>
<td>{{ a[4] }}</td><td>{{ a[5] }}</td><td>{{ a[6] }}</td><td>{{ a[7] }}</td>
<td><span class="badge {{ 'p' if a[9]=='Pending' else '' }}">{{ a[9] }}</span></td></tr>
{% else %}<tr><td colspan="8" class="te">No appointments yet.</td></tr>{% endfor %}</table></div>

<h2>💰 Estimates</h2>
<div class="table-wrap"><table><tr><th>ID</th><th>Customer</th><th>Service</th><th>Type</th><th>Details</th><th>Sqft</th><th>Estimate</th><th>Date</th></tr>
{% for e in ests|reverse %}<tr><td>{{ e[0] }}</td><td>{{ e[2] }}<br><span class="te">{{ e[1] }}</span></td>
<td>{{ e[3] }}</td><td>{{ e[4] }}</td><td style="max-width:280px;white-space:normal">{{ e[5] }}</td>
<td>{{ e[6] }}</td><td><b>₹{{ e[7]|inr }}</b></td><td>{{ e[8] }}</td></tr>
{% else %}<tr><td colspan="8" class="te">No estimates yet.</td></tr>{% endfor %}</table></div>

<h2>👥 Customers</h2>
<div class="table-wrap"><table><tr><th>ID</th><th>Name</th><th>Email</th><th>Phone</th><th>Signed Up</th></tr>
{% for u in users|reverse %}<tr><td>{{ u[0] }}</td><td>{{ u[1] }}</td><td>{{ u[2] }}</td>
<td><a href="tel:{{ u[3] }}">{{ u[3] }}</a></td><td>{{ u[5] }}</td></tr>
{% else %}<tr><td colspan="5" class="te">No customers yet.</td></tr>{% endfor %}</table></div>

<h2 id="security">🛡️ Security</h2>
<div class="card">
  <p class="te" style="margin-bottom:10px">
    {% if alerts_on %}Email alerts are <b>ON</b> — suspicious activity is sent to {{ alert_to }}.
    {% else %}Email alerts are <b>OFF</b>. Set ALERT_EMAIL, SMTP_USER and SMTP_PASS in
    Render → Settings → Environment to start receiving them.{% endif %}
  </p>
  <form method="post" action="{{ url_for('test_alert') }}" style="margin:0">
    <input type="hidden" name="_csrf" value="{{ csrf_token() }}">
    <button class="btn btn-outline">Send me a test alert</button>
  </form>
</div>
<div class="table-wrap"><table><tr><th>When (IST)</th><th>Severity</th><th>Event</th><th>From IP</th><th>Detail</th></tr>
{% for e in events %}<tr>
<td>{{ e.at|ist }}</td>
<td><span class="badge {{ 'p' if e.severity in ('high','critical') else '' }}">{{ e.severity }}</span></td>
<td>{{ e.kind }}</td><td>{{ e.ip }}</td>
<td style="max-width:320px;white-space:normal">{{ e.detail }}</td></tr>
{% else %}<tr><td colspan="5" class="te">Nothing suspicious recorded since the last restart.</td></tr>{% endfor %}</table></div>
</div>""", users=users, ests=ests, appts=appts, tv=total_value, events=events,
           alerts_on=security.alerts_configured(), alert_to=security.ALERT_EMAIL)

@app.route("/admin/test-alert", methods=["POST"])
@admin_required
def test_alert():
    if security.send_test_alert():
        flash(f"Test alert sent to {security.ALERT_EMAIL}. Check your inbox (and spam).")
    else:
        flash("Email is not configured yet — set ALERT_EMAIL, SMTP_USER and SMTP_PASS "
              "in Render → Settings → Environment.")
    return redirect(url_for("admin") + "#security")

@app.route("/admin/download")
@admin_required
def download_excel():
    return send_file(EXCEL_FILE, as_attachment=True,
                     download_name=f"gvm_infra_developers_data_{datetime.now():%Y%m%d}.xlsx")

# ------------------------------- MAIN ----------------------------------
if __name__ == "__main__":
    # The dev server speaks plain HTTP, so Secure cookies would never be sent
    # back and login would silently fail. Production (gunicorn on Render) keeps
    # them on; see SECURE_COOKIES in security.py.
    app.config["SESSION_COOKIE_SECURE"] = False
    print("Local dev: Secure cookies disabled (HTTP). Production keeps them ON.")
    app.run(host="0.0.0.0", port=5000, debug=False)
