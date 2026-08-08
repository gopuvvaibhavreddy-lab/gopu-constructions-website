"""Automated test suite for the GVM Infra Developers website. Run: python3 test_site.py"""
import os, sys, re, tempfile

os.chdir(tempfile.mkdtemp())  # fresh dir so a fresh data.xlsx is created
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ["SECURE_COOKIES"] = "0"   # the test client speaks plain HTTP
import importlib, app as appmod, security, mailer
appmod.EXCEL_FILE = os.path.join(os.getcwd(), "data.xlsx")
appmod.init_excel()
app = appmod.app
app.config["TESTING"] = True
app.config["SESSION_COOKIE_SECURE"] = False

# Capture mail instead of sending it, and keep the log out of test output.
SENT = []                       # security alert events
MAILED = []                     # every EmailMessage handed to the transport
security.ALERT_SENDER = SENT.append
mailer.TRANSPORT = MAILED.append
import logging
logging.getLogger("gvm.security").setLevel(logging.CRITICAL)
logging.getLogger("gvm.mailer").setLevel(logging.CRITICAL)

def mail_to(addr):
    return [m for m in MAILED if m["To"] == addr]

P = F = 0
def t(name, got, want):
    global P, F
    if got == want: P += 1
    else: F += 1; print(f"  FAIL {name}: got {got!r}, want {want!r}")

c = app.test_client()

def tok():
    """Current session's CSRF token, scraped from a page with a form on it."""
    m = re.search(r'name="_csrf" value="([^"]+)"', c.get("/signup").data.decode())
    return m.group(1) if m else ""

def post(path, data=None, **kw):
    """POST with a valid CSRF token, the way a real browser would."""
    d = dict(data or {})
    d["_csrf"] = tok()
    return c.post(path, data=d, **kw)

def kinds():
    return [e["kind"] for e in security.recent_events(60)]

# ---- public pages & guards
home_html = c.get("/").data.decode()
t("home", c.get("/").status_code, 200)
t("home-has-testimonials-section", 'id="testimonials"' in home_html, True)
t("home-testimonials-count", home_html.count('class="card testimonial"'), 5)
t("404", c.get("/nope").status_code, 404)
t("signup-page", c.get("/signup").status_code, 200)
for p in ["/dashboard", "/renovation", "/build-new", "/appointment"]:
    t(f"guard{p}", c.get(p).status_code, 302)
t("guard-admin", c.get("/admin").status_code, 302)
t("guard-excel", c.get("/admin/download").status_code, 302)

# ---- signup validation
t("signup-short-pw", post("/signup", data=dict(name="A", email="a@b.com", phone="1234567890", password="123")).status_code, 302)
t("signup-bad-email", post("/signup", data=dict(name="A", email="nope", phone="1234567890", password="123456")).status_code, 302)
t("signup-empty-name", post("/signup", data=dict(name=" ", email="a@b.com", phone="1234567890", password="123456")).status_code, 302)
t("signup-missing-fields", post("/signup", data={}).status_code, 302)

# ---- good signup / duplicate / login
r = post("/signup", data=dict(name="Ravi Kumar", email="ravi@test.com", phone="9876543210", password="secret12"))
t("signup-ok", r.status_code, 302)
t("dashboard-after-signup", c.get("/dashboard").status_code, 200)
c.get("/logout")
t("dup-signup-case-insensitive", post("/signup", data=dict(name="R", email="RAVI@test.com", phone="9876543210", password="secret12"), follow_redirects=True).request.path, "/login")
t("login-wrong-pw", b"Invalid" in post("/login", data=dict(email="ravi@test.com", password="WRONG")).data, True)
t("login-unknown-user", b"Invalid" in post("/login", data=dict(email="ghost@x.com", password="x")).data, True)
t("login-ok", post("/login", data=dict(email="Ravi@Test.com", password="secret12")).status_code, 302)

# ---- renovation pages
t("reno-choose", c.get("/renovation").status_code, 200)
t("reno-res", c.get("/renovation?type=residential").status_code, 200)
t("reno-com", c.get("/renovation?type=commercial").status_code, 200)
t("reno-badtype-shows-chooser", c.get("/renovation?type=hack").status_code, 200)
html = c.get("/renovation?type=residential").data.decode()
t("form-keys-no-spaces", 'name="sqft_Full_House"' in html, True)
t("telugu-present", "పడక గది" in html, True)
t("html-not-escaped", "&lt;tr&gt;" not in html and '<input type="number"' in html, True)
t("home-html-not-escaped", "&lt;div" not in c.get("/").data.decode(), True)
t("live-calc-script-present", "function calc()" in html, True)
t("reno-table-wrapped", 'class="table-wrap"' in html, True)

build_html = c.get("/build-new?type=residential").data.decode()
t("build-new-page", c.get("/build-new?type=commercial").status_code, 200)
t("build-new-table-wrapped", 'class="table-wrap"' in build_html, True)

# ---- estimates
r = post("/save-estimate", data=dict(service="Renovation", ptype="residential",
    sqft_Bedroom="100", sqft_Full_House="500"), follow_redirects=False)
t("reno-estimate-status", r.status_code, 200)
t("reno-estimate-total-730000", "₹7,30,000" in r.data.decode(), True)  # 100*1050 + 500*1250
t("est-garbage-input", post("/save-estimate", data=dict(service="Renovation", ptype="residential", sqft_Bedroom="abc")).status_code, 302)
t("est-negative", post("/save-estimate", data=dict(service="Renovation", ptype="residential", sqft_Bedroom="-50")).status_code, 302)
t("est-all-empty", post("/save-estimate", data=dict(service="Renovation", ptype="residential")).status_code, 302)
t("est-bad-ptype", post("/save-estimate", data=dict(service="Renovation", ptype="HACK", sqft_Bedroom="10")).status_code, 302)
t("est-bad-service", post("/save-estimate", data=dict(service="HACK", ptype="residential")).status_code, 302)
t("est-no-fields", post("/save-estimate", data={}).status_code, 302)
r = post("/save-estimate", data=dict(service="Build New", ptype="residential", sqft_total="1500", tier="Premium"))
t("build-estimate-3900000", "₹39,00,000" in r.data.decode(), True)  # 1500*2600
r = post("/save-estimate", data=dict(service="Build New", ptype="commercial", sqft_total="1000", tier="HACKTIER"))
t("build-bad-tier-falls-back-standard", "₹22,00,000" in r.data.decode(), True)  # 1000*2200
t("build-huge-number", post("/save-estimate", data=dict(service="Build New", ptype="residential", sqft_total="999999999", tier="Basic")).status_code, 200)

# ---- appointments
t("appt-page", c.get("/appointment").status_code, 200)
r = post("/appointment", data=dict(name="Ravi", phone="9876543210", service="Renovation - House",
    date="2026-07-15", time="Morning (9 AM - 12 PM)", location="Warangal", notes="near temple"))
t("appt-ok", r.status_code, 200)
t("appt-whatsapp-link", "wa.me" in r.data.decode(), True)
t("appt-missing-fields", post("/appointment", data=dict(name="", phone="", date="", location="")).status_code, 302)
# XSS attempt stored
post("/appointment", data=dict(name="<script>alert(1)</script>", phone="1234567890",
    service="Other / Consultation", date="2026-07-20", time="Morning (9 AM - 12 PM)", location="Hyd"))

# ---- admin
c.get("/logout")
t("admin-wrong", b"Wrong" in post("/admin-login", data=dict(username="admin", password="nope")).data, True)
t("admin-ok", post("/admin-login", data=dict(username="admin", password=appmod.ADMIN_PASSWORD)).status_code, 302)
admin_html = c.get("/admin").data.decode()
t("admin-page", c.get("/admin").status_code, 200)
t("admin-shows-customer", "Ravi Kumar" in admin_html, True)
t("admin-shows-estimate", "₹7,30,000" in admin_html, True)
t("admin-shows-appt", "Warangal" in admin_html, True)
t("xss-escaped", "<script>alert" not in admin_html, True)
t("admin-tables-wrapped", admin_html.count('class="table-wrap"'), 4)
t("excel-download", c.get("/admin/download").status_code, 200)

# ---- excel integrity
from openpyxl import load_workbook
wb = load_workbook(appmod.EXCEL_FILE)
t("excel-sheets", wb.sheetnames, ["Users", "Estimates", "Appointments"])
users = list(wb["Users"].iter_rows(min_row=2, values_only=True))
t("excel-1-user", len(users), 1)
t("excel-pw-hashed", users[0][4].startswith(("scrypt:", "pbkdf2:")), True)
ests = list(wb["Estimates"].iter_rows(min_row=2, values_only=True))
t("excel-estimates-count", len(ests), 4)
appts = list(wb["Appointments"].iter_rows(min_row=2, values_only=True))
t("excel-appts-count", len(appts), 2)
t("excel-appt-status-pending", appts[0][9], "Pending")

# ---- excel self-heal after deletion
os.remove(appmod.EXCEL_FILE)
t("self-heal-admin", c.get("/admin").status_code, 200)
c.get("/logout")
t("self-heal-signup", post("/signup", data=dict(name="New", email="new@x.com", phone="1112223334", password="pass1234")).status_code, 302)

# ---- inr formatting
t("inr-0", appmod.inr(0), "0")
t("inr-999", appmod.inr(999), "999")
t("inr-lakh", appmod.inr(123456), "1,23,456")
t("inr-crore", appmod.inr(12345678), "1,23,45,678")
t("inr-none", appmod.inr(None), "0")

# ---- security: headers
h = c.get("/").headers
t("hdr-nosniff", h.get("X-Content-Type-Options"), "nosniff")
t("hdr-frameopts", h.get("X-Frame-Options"), "DENY")
t("hdr-referrer", h.get("Referrer-Policy"), "strict-origin-when-cross-origin")
t("hdr-csp-present", "default-src 'self'" in h.get("Content-Security-Policy", ""), True)
t("hdr-csp-no-frame", "frame-ancestors 'none'" in h.get("Content-Security-Policy", ""), True)
t("hdr-csp-allows-google-fonts", "fonts.googleapis.com" in h.get("Content-Security-Policy", ""), True)
t("hdr-permissions-policy", "camera=()" in h.get("Permissions-Policy", ""), True)

# ---- security: CSRF
t("csrf-field-on-login-form", 'name="_csrf"' in c.get("/login").data.decode(), True)
t("csrf-missing-token-rejected", c.post("/login", data=dict(email="a@b.com", password="x")).status_code, 400)
t("csrf-wrong-token-rejected", c.post("/login", data=dict(email="a@b.com", password="x", _csrf="bogus")).status_code, 400)
t("csrf-failure-recorded", "csrf_failure" in kinds(), True)
t("csrf-valid-token-accepted", post("/login", data=dict(email="a@b.com", password="x")).status_code, 200)

# ---- security: spreadsheet formula injection
t("formula-neutralised", appmod.security.spreadsheet_safe('=HYPERLINK("http://evil","x")')[0], "'")
t("formula-plus", appmod.security.spreadsheet_safe("+1+1")[0], "'")
t("formula-at", appmod.security.spreadsheet_safe("@SUM(A1)")[0], "'")
t("normal-text-untouched", appmod.security.spreadsheet_safe("Ravi Kumar"), "Ravi Kumar")
t("non-string-untouched", appmod.security.spreadsheet_safe(1500), 1500)
c.get("/logout")
post("/signup", data=dict(name="=cmd|'/c calc'!A1", email="formula@x.com",
                          phone="9998887770", password="passw0rd1"))
wb2 = load_workbook(appmod.EXCEL_FILE)
formula_user = [r for r in wb2["Users"].iter_rows(min_row=2, values_only=True)
                if r[2] == "formula@x.com"]
t("formula-name-stored-inert", formula_user[0][1].startswith("'="), True)

# ---- security: payload detection (alert-only, never blocks a customer)
before = len(kinds())
c.get("/renovation?type=residential' UNION SELECT password FROM users--")
t("sqli-detected", "sql_injection" in kinds(), True)
t("xss-detected-from-earlier-appointment", "xss_attempt" in kinds(), True)
t("payload-detection-does-not-block", c.get("/renovation?type=residential").status_code, 200)

# ---- security: scanner probes
t("probe-404", c.get("/.git/HEAD").status_code, 404)
t("probe-recorded", "probe_path" in kinds(), True)
for i in range(16):
    c.get("/wp-admin/install.php")
t("sustained-scan-escalates", "vuln_scan" in kinds(), True)

# ---- security: admin brute-force lockout
c.get("/logout")
for i in range(appmod.security.ADMIN_FAIL_LIMIT[0]):
    post("/admin-login", data=dict(username="admin", password=f"guess{i}"))
t("admin-bruteforce-alert", "admin_bruteforce" in kinds(), True)
t("admin-locked-out", appmod.security.is_locked("adminfail", "127.0.0.1"), True)
t("admin-locked-rejects-good-password",
  post("/admin-login", data=dict(username="admin", password=appmod.ADMIN_PASSWORD),
       follow_redirects=True).request.path, "/admin-login")
appmod.security._locks_until.clear()
t("admin-works-after-lock-expires",
  post("/admin-login", data=dict(username="admin", password=appmod.ADMIN_PASSWORD)).status_code, 302)
t("admin-login-success-alerted", "admin_login_ok" in kinds(), True)

# ---- security: alerts were queued, not emailed
import time
time.sleep(0.4)  # daemon threads deliver to the fake sender
t("alerts-were-sent-to-fake-sender", len(SENT) > 0, True)
t("alert-has-ip-and-kind", all(("ip" in e and "kind" in e) for e in SENT), True)
t("no-low-severity-emails", all(e["severity"] != "low" for e in SENT), True)
t("alert-body-renders", "From IP:" in appmod.security._body(SENT[0]), True)
t("alerts-not-configured-in-tests", appmod.security.alerts_configured(), False)

# ---- company email on the site
home2 = c.get("/").data.decode()
t("footer-shows-company-email", "vaibhavreddy@gvminfradevelopers.com" in home2, True)
t("footer-email-is-mailto", 'href="mailto:vaibhavreddy@gvminfradevelopers.com"' in home2, True)
t("footer-phone-is-tel", 'href="tel:+918332899003"' in home2, True)
t("company-email-on-every-page", "vaibhavreddy@gvminfradevelopers.com" in c.get("/login").data.decode(), True)
# hidden rather than shown broken if it is ever blanked out
_saved = appmod.COMPANY_EMAIL
appmod.COMPANY_EMAIL = ""
t("blank-company-email-hides-link", "mailto:" not in c.get("/").data.decode(), True)
appmod.COMPANY_EMAIL = _saved

# ---- welcome email on signup
import time
c.get("/logout")
MAILED.clear()
# earlier tests already created accounts; start these from a clean per-IP count
security._hits.clear()
r = post("/signup", data=dict(name="Anjali Rao", email="anjali@test.com",
                              phone="9123456780", password="welcome123"))
t("welcome-signup-ok", r.status_code, 302)
time.sleep(0.5)                       # daemon thread hands off to the fake transport
wm = mail_to("anjali@test.com")
t("welcome-email-sent", len(wm), 1)
t("welcome-subject", "Welcome to GVM Infra Developers" in wm[0]["Subject"], True)
t("welcome-reply-to-company", wm[0]["Reply-To"], "vaibhavreddy@gvminfradevelopers.com")
t("welcome-from-has-company-name", "GVM Infra Developers" in wm[0]["From"], True)
t("welcome-is-multipart-text-and-html", wm[0].is_multipart(), True)
_bodies = "".join(p.get_content() for p in wm[0].walk() if p.get_content_type().startswith("text/"))
t("welcome-greets-by-name", "Anjali Rao" in _bodies, True)
t("welcome-has-telugu", "మీకు స్వాగతం" in _bodies, True)
t("welcome-has-site-link", "gvminfradevelopers.com" in _bodies, True)
t("welcome-has-whatsapp", "wa.me/918332899003" in _bodies, True)
t("welcome-shows-company-email", "vaibhavreddy@gvminfradevelopers.com" in _bodies, True)

# a name containing HTML must not land unescaped in the HTML part
MAILED.clear()
c.get("/logout")
post("/signup", data=dict(name="<b>Bold</b> Ravi", email="htmlname@test.com",
                          phone="9123456781", password="welcome123"))
time.sleep(0.5)
_html = [p.get_content() for p in mail_to("htmlname@test.com")[0].walk()
         if p.get_content_type() == "text/html"][0]
t("welcome-html-escapes-name", "<b>Bold</b>" not in _html, True)
t("welcome-html-keeps-name-text", "&lt;b&gt;Bold&lt;/b&gt;" in _html, True)

# signup must still succeed when mail is broken
def _boom(msg):
    raise RuntimeError("SMTP down")
mailer.TRANSPORT = _boom
c.get("/logout")
t("signup-survives-mail-failure",
  post("/signup", data=dict(name="Mail Down", email="maildown@test.com",
                            phone="9123456782", password="welcome123")).status_code, 302)
time.sleep(0.4)
t("user-saved-despite-mail-failure", appmod.find_user("maildown@test.com") is not None, True)
mailer.TRANSPORT = MAILED.append

# no welcome mail for a rejected signup
MAILED.clear()
c.get("/logout")
post("/signup", data=dict(name="Bad", email="nope", phone="1", password="short"))
time.sleep(0.3)
t("no-welcome-for-invalid-signup", len(MAILED), 0)

# ---- signup rate limit permits exactly SIGNUP_LIMIT accounts per window
security._hits.clear()
LIM = appmod.security.SIGNUP_LIMIT[0]
made = 0
for i in range(LIM + 2):
    c.get("/logout")
    post("/signup", data=dict(name=f"Rate {i}", email=f"rate{i}@test.com",
                              phone="9000000000", password="welcome123"))
    if appmod.find_user(f"rate{i}@test.com"):
        made += 1
t("signup-limit-allows-exactly-limit", made, LIM)
t("signup-abuse-alert-raised", "signup_abuse" in kinds(), True)
security._hits.clear()

# ---- no SyntaxWarnings (an invalid \escape in a template string is a real bug:
#      it silently changes the HTML and breaks in a future Python)
import warnings, py_compile, tempfile as _tf
_src = os.path.join(os.path.dirname(os.path.abspath(appmod.__file__)), "app.py")
with warnings.catch_warnings(record=True) as caught:
    warnings.simplefilter("always")
    py_compile.compile(_src, cfile=_tf.mktemp(), doraise=True)
t("app-compiles-without-warnings", [str(w.message) for w in caught], [])
signup_html = c.get("/signup").data.decode()
t("phone-pattern-escapes-hyphen", r'pattern="[0-9+ \-]{10,15}"' in signup_html, True)

# ---- security: admin dashboard shows the log
# the signup tests above logged out, so sign back in as admin first
security._locks_until.clear()
post("/admin-login", data=dict(username="admin", password=appmod.ADMIN_PASSWORD))
admin_html2 = c.get("/admin").data.decode()
t("admin-reauth-ok", c.get("/admin").status_code, 200)
t("admin-has-security-section", 'id="security"' in admin_html2, True)
t("admin-lists-events", "admin_bruteforce" in admin_html2, True)
t("admin-shows-alert-status", "Email alerts are" in admin_html2, True)

print(f"\n{'='*40}\nRESULT: {P} passed, {F} failed\n{'='*40}")
sys.exit(1 if F else 0)
